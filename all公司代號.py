a=open('上市公司代號.txt','r',encoding="utf-8")
b=a.read()
from bs4 import  BeautifulSoup
soup=BeautifulSoup(b,'html.parser')
l='td'
lm=soup.select(l)
list=[]
#print(len(i))
for i in range(0,6730,7):
    list.append(lm[i].string.get_text())
代號=[]
公司名稱=[]
for i in range(len(list)):
    代號.append(list[i][:4])
    公司名稱.append(list[i][5:])
a.close()
#############################################
aa=open('上櫃公司代號.txt','r',encoding="utf-8")
bb=aa.read()

from bs4 import  BeautifulSoup
soup=BeautifulSoup(bb,'html.parser')
ll='td'
lmm=soup.select(l)
listt=[]
#print(len(i))
for i in range(0,5572,7):
    listt.append(lmm[i].string.get_text())

代號1=[]
公司名稱1=[]
for i in range(len(listt)):
    代號1.append(listt[i][:4])
    公司名稱1.append(listt[i][5:])
aa.close()
#################################################
aaa=open('興櫃公司代號.txt','r',encoding="utf-8")
bbb=aaa.read()

from bs4 import  BeautifulSoup
soup=BeautifulSoup(bbb,'html.parser')
lll='td'
lmmm=soup.select(l)
listtt=[]
#print(len(i))
for i in range(0,2044,7):
    listtt.append(lmmm[i].string.get_text())

代號2=[]
公司名稱2=[]
for i in range(len(listtt)):
    代號2.append(listtt[i][:4])
    公司名稱2.append(listtt[i][5:])
aaa.close()







from openpyxl import Workbook,load_workbook
wb=Workbook()#創建新表格檔案
ws=wb.active#叫出當前執行工作區域
#print(ws)#印出當前執行的工作區域名稱

ws.title="上市公司"#更改當前執行工作區域名稱
#print(ws)#印出當前執行的工作區域名稱
wb.create_sheet("上櫃公司")#新增工作區域
wb.create_sheet("興櫃公司")#新增工作區域
wb.create_sheet("所有上市櫃公司")#新增工作區域
sheet_names = wb.sheetnames   # 返回一个列表
ws4 = wb[sheet_names[3]] 
ws3 = wb[sheet_names[2]] 
ws2 = wb[sheet_names[1]] 
ws1 = wb[sheet_names[0]] 

for i in range(len(代號)):
    ws1[f'A{i+1}']=int(代號[i])
    ws1[f'B{i+1}']=公司名稱[i]
    ws4[f'A{i+1}']=int(代號[i])
    ws4[f'B{i+1}']=公司名稱[i]



for i in range(len(代號1)):
    ws2[f'A{i+1}']=int(代號1[i])
    ws2[f'B{i+1}']=公司名稱1[i]
    ws4[f'D{i+1}']=int(代號1[i])
    ws4[f'E{i+1}']=公司名稱1[i]

for i in range(len(代號2)):
    ws3[f'A{i+1}']=int(代號2[i])
    ws3[f'B{i+1}']=公司名稱2[i]
    ws4[f'G{i+1}']=int(代號2[i])
    ws4[f'H{i+1}']=公司名稱2[i]





#儲存
wb.save(r"上市櫃公司代號.xlsx")#儲存新表格